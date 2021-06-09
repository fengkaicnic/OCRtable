# -*- coding: utf-8 -*-
import os, openpyxl
import pdb
from openpyxl.styles import Border,Side
from XlsBorder import XLSBorder
#path = 'D:/ORC'
#os.chdir(path)

v_lines = [[23, 98, 23, 523],
[261, 412, 261, 523],
[451, 98, 451, 224],
[451, 239, 451, 292],
[451, 332, 451, 412],
[451, 448, 451, 523],
[578, 131, 578, 292],
[578, 332, 578, 412],
[578, 448, 578, 523],
[704, 98, 704, 523]]

h_lines = [[23, 98, 705, 98],
[451, 131, 705, 131],
[578, 171, 705, 171],
[23, 212, 705, 212],
[451, 252, 705, 252],
[23, 292, 705, 292],
[23, 332, 705, 332],
[451, 372, 705, 372],
[23, 412, 705, 412],
[23, 448, 705, 448],
[23, 487, 705, 487],
[23, 522, 705, 522]]

rect_lst = [((451, 98), (704, 98), (451, 131), (704, 131)),
((23, 98), (451, 98), (23, 212), (451, 212)),
((578, 131), (704, 131), (578, 171), (704, 171)),
((451, 131), (578, 131), (451, 212), (578, 212)),
((578, 171), (704, 171), (578, 212), (704, 212)),
((451, 212), (578, 212), (451, 252), (578, 252)),
((578, 212), (704, 212), (578, 252), (704, 252)),
((23, 212), (451, 212), (23, 292), (451, 292)),
((451, 252), (578, 252), (451, 292), (578, 292)),
((578, 252), (704, 252), (578, 292), (704, 292)),
((23, 292), (704, 292), (23, 332), (704, 332)),
((451, 332), (578, 332), (451, 372), (578, 372)),
((578, 332), (704, 332), (578, 372), (704, 372)),
((23, 332), (451, 332), (23, 412), (451, 412)),
((451, 372), (578, 372), (451, 412), (578, 412)),
((578, 372), (704, 372), (578, 412), (704, 412)),
((23, 412), (261, 412), (23, 448), (261, 448)),
((261, 412), (704, 412), (261, 448), (704, 448)),
((23, 448), (261, 448), (23, 487), (261, 487)),
((261, 448), (451, 448), (261, 487), (451, 487)),
((451, 448), (578, 448), (451, 487), (578, 487)),
((578, 448), (704, 448), (578, 487), (704, 487)),
((23, 487), (261, 487), (23, 522), (261, 522)),
((261, 487), (451, 487), (261, 522), (451, 522)),
((451, 487), (578, 487), (451, 522), (578, 522)),
((578, 487), (704, 487), (578, 522), (704, 522))]

codct = {1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G', 8:'H',
         9:'I', 10:'J', 11:'K', 12:'L', 13:'M', 14:'N'}
dict2 = {value: key for key, value in codct.items()}

horxl_lst = []
for vet in v_lines:
    if len(horxl_lst):
        if horxl_lst[-1] != vet[0]:
            horxl_lst.append(vet[0])
    else:
        horxl_lst.append(vet[0])

verxl_lst = []
for vet in h_lines:
    if len(verxl_lst):
        if verxl_lst[-1] != vet[1]:
            verxl_lst.append(vet[1])
    else:
        verxl_lst.append(vet[1])

print (horxl_lst)
print (verxl_lst)

merlst = []
#(23, 292), (451, 292), (23, 332), (451, 332)
for rect in rect_lst:
    a = rect[0]
    b = rect[-1]
    hstart, hend = a[0], b[0]
    vstart, vend = a[1], b[1]
    cols = codct.get(horxl_lst.index(hstart)+1, None)
    cole = codct.get(horxl_lst.index(hend), None)
    # pdb.set_trace()
    rows = verxl_lst.index(vstart)+1
    rowe = verxl_lst.index(vend)
    print(cols+str(rows)+':'+cole+str(rowe))
    merlst.append(cols+str(rows)+':'+cole+str(rowe))

print(len(merlst), merlst)

pdb.set_trace()
wb = openpyxl.Workbook()
# ws.column_dimensions['A'].width = 20.0
# ws.row_dimensions[1].height = 40
sheet = wb.get_sheet_by_name('Sheet')

for cell in merlst:
    if cell[0] == cell[1]:
        continue
    # col = dict2.get(cell[0])
    col = cell.split(':')[0][0]
    row = int(cell.split(':')[0][1:])
    sheet.column_dimensions[col].width = 20.0
    sheet.row_dimensions[row].height = 30
    # top_left_cell = sheet[cell[0]]
    # top_left_cell.value = "My Cell"
    # sheet.cell(row, col).value = cell
    # pdb.set_trace()
    print (cell)
    sheet[cell[:2]].value = cell
    # sheet.merge_cells(cell)
xlsBorder = XLSBorder()
# xlsBorder.format_border('A1:D11', sheet)
for cell in merlst:
    # print (cell)
    xlsBorder.set_solid_border(cell, sheet)
    sheet.merge_cells(cell)

wb.save('merge1.xlsx')

# sheet.merge_cells('B2:D4')
# sheet.cell(row=2, column=2).value = 'merge 9 cell'
# sheet.merge_cells('C5:D6')
# sheet['C5'] = 'merge 4 cell'
# wb.save('合并单元格.xlsx')

