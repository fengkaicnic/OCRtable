# -*- coding: utf-8 -*-

import sys
import openpyxl
import pdb
import MySQLdb as mysql

# mysql信息
host = '127.0.0.1'
port = 3306
user = 'root'
passwd = '123456'
db = 'exceldb'
charset = 'utf8'

# 连接mysql，获取cursor
conn = mysql.connect(host=host, port=port, user=user, passwd=passwd, db=db, charset=charset)
cursor = conn.cursor()

table = 'testexc'
count = cursor.execute("select * from {table};".format(table=table))
print(count)

# 重置游标的位置
cursor.scroll(0, mode='absolute')
# 拿到该条SQL所有结果
results = cursor.fetchall()
print(results)

# 拿到该表里面的字段名称
fields = cursor.description
print(fields)

ws = openpyxl.Workbook()
sheet = ws.get_sheet_by_name(ws.sheetnames[0])

# 写上字段信息
for field in range(1, len(fields)):
   sheet.cell(row=1, column=field).value = fields[field][0]

pdb.set_trace()

# 获取并写入数据段信息
# row = 1
# col = 0
for rownm in range(2, len(results) + 1):
    # pdb.set_trace()
    for col in range(1, len(fields)):
        value = results[rownm - 2][col]
        if not value:
           value = ''
        sheet.cell(row=rownm, column=col).value = value

ws.save(r'./{db}_{table}.xlsx'.format(db=db, table=table))


