# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import MySQLdb
import pdb

# path = sys.argv[1]
path = '附表4.仓储财务报表（出库）.xlsx'

wb = load_workbook(path)
print(wb.sheetnames)
ws = wb.get_sheet_by_name(wb.sheetnames[0])
tablename = ws['A1'].value
print(tablename)
columnlst = []
rownm = 2
insqls = ['insert into testexc(']
for i in range(13):
    # print(ws.cell(row=rownm, column=i+1).value)
    columnlst.append(ws.cell(row=rownm, column=i+1).value)
    insqls.append(ws.cell(row=rownm, column=i+1).value + ',')
insqls[-1] = insqls[-1][:-1]
insqls.append(') values (')

conn= MySQLdb.connect(
        host='localhost',
        port = 3306,
        user='root',
        passwd='123456',
        db ='exceldb',
        charset='utf8'
        )
cur = conn.cursor()
#'create table testexc(id int unsigned not null primary key auto_increment,
#  序号 varchar(30),材料类型 varchar(30),材料名称 varchar(30),规格 varchar(30),
#  单位 varchar(30),数量 varchar(30),  币种 varchar(30),单价 varchar(30),
#  总价 varchar(30),领用部门 varchar(30),领料人 varchar(30),出库日期 varchar(30), 备注 varchar(30)'

sql = ["create table testexc(id int unsigned not null primary key auto_increment"]
for col in columnlst:
    sql.append(','+col+ ' varchar(30)')
# cur.execute("create table student(id int ,name varchar(20),class varchar(30),age varchar(10))")
# pdb.set_trace()
cur.execute(''.join(sql) + ')')
import copy
for ro in range(3, 19):
    insql = copy.deepcopy(insqls)
    for cl in range(1, 14):
        insql.append('"'+str(ws.cell(row=ro, column=cl).value)+'",')
    insql[-1] = insql[-1][:-1]
    # pdb.set_trace()
    cur.execute(''.join(insql)+')')

cur.close()
conn.commit()
conn.close()
# print(''.join(insql))


