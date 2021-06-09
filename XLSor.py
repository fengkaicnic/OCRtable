# -*- coding: utf-8 -*-
import os, openpyxl
import pdb
from openpyxl.styles import Border,Side
from XlsBorder import XLSBorder
#path = 'D:/ORC'
#os.chdir(path)

class XLsor(object):
    def __init__(self, rect_lst, vlines, hlines, rectText):

        # self.gray = gray
        self.vlines = vlines
        self.hlines = hlines
        self.rect_lst = rect_lst
        self.rectText = rectText
        self.codct = {1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G', 8:'H',
                 9:'I', 10:'J', 11:'K', 12:'L', 13:'M', 14:'N'}
        self.rectCell = {}

    def genXls(self, hline):

        horxl_lst = []
        #用竖线判断横格有多少，就是有几行
        for vet in self.vlines:
            if len(horxl_lst):
                if horxl_lst[-1] != vet[0]:
                    horxl_lst.append(vet[0])
            else:
                horxl_lst.append(vet[0])

        verxl_lst = []
        #用横线判断有几列， 每条横线的纵坐标都是一样的，所以verxl_lst是纵坐标列表
        verxl_lst.append(hline[1][1])
        for vet in self.hlines:
            if len(verxl_lst):
                if verxl_lst[-1] != vet[1]:
                    verxl_lst.append(vet[1])
            else:
                verxl_lst.append(vet[1])

        print (horxl_lst)
        print (verxl_lst)

        merlst = []
        #(23, 292), (451, 292), (23, 332), (451, 332)
        pdb.set_trace()
        for rect in self.rect_lst:
            a = rect[0] #a是方块的第一个点
            b = rect[-1]#b是方块的最后一个点，也就是右下的点
            hstart, hend = a[0], b[0]#hstart和hend分别代表了方块的两个横坐标
            vstart, vend = a[1], b[1]#vstart和vend代表了方块的连个纵坐标
            cols = self.codct.get(horxl_lst.index(hstart)+1, None)
            cole = self.codct.get(horxl_lst.index(hend), None)
            # pdb.set_trace()
            rows = verxl_lst.index(vstart)+1
            rowe = verxl_lst.index(vend)
            print(cols+str(rows)+':'+cole+str(rowe))
            merlst.append(cols+str(rows)+':'+cole+str(rowe))
            self.rectCell[rect] = merlst[-1]

        print(len(merlst), merlst)

        # pdb.set_trace()
        wb = openpyxl.Workbook()
        sheet = wb.get_sheet_by_name('Sheet')
        # for cell in merlst:
        pdb.set_trace()
        for rect, cell in self.rectCell.items():
            print(rect, cell)
            col = cell.split(':')[0][0]
            row = int(cell.split(':')[0][1:])
            sheet.column_dimensions[col].width = 20.0
            sheet.row_dimensions[row].height = 30
            if self.rectText.get(rect):
                nm = cell.split(':')[0]
                sheet[nm].value = ''.join(self.rectText[rect])
        xlsBorder = XLSBorder()
        for rect, cell in self.rectCell.items():
            if cell[0] == cell[1]:
                continue
            xlsBorder.set_solid_border(cell, sheet)
            sheet.merge_cells(cell)


        wb.save('mergetest.xlsx')

# sheet.merge_cells('B2:D4')
# sheet.cell(row=2, column=2).value = 'merge 9 cell'
# sheet.merge_cells('C5:D6')
# sheet['C5'] = 'merge 4 cell'
# wb.save('合并单元格.xlsx')

